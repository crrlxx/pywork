#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/1/20 15:48
# @Author  : crrlxx
# @Site    : 
# @File    : bankdyy.py
# @Software: PyCharm

import time
import datetime
import requests
import sys
import lxml.etree as etree
import os
import json
import urllib
# import urllib.request
import openpyxl.workbook as workbook
from openpyxl import load_workbook

# reload(sys)
# sys.setdefaultencoding("utf-8")

days = []
for i in range(7):
    days.insert(0, str(datetime.date.today() - datetime.timedelta(i)))

homepath = os.getcwd()+os.sep+"excel"+os.sep
print("Checking date:", days[0], days[6])
print("Working directory:", homepath)
wb_out = workbook.Workbook()
ws_out = wb_out.get_sheet_by_name("Sheet")
i_row = 1
i_col = 1
shibor_step = ['O/N', '1W','2W','1M','3M','6M','9M','1Y']
all_code = {'Shibor': 'Shibor', 'CYCC000': '国债', 'CYCC021': '政策性金融债（国开）',
             'CYCC82A': '中期票据AAA+', 'CYCC82B': '中期票据AAA', 'CYCC82C': '中期票据AAA-', 'CYCC82D': '中期票据AA+',
             'CYCC81A': '短期融资券AAA+', 'CYCC81B': '短期融资券AAA', 'CYCC81C': '短期融资券AAA-', 'CYCC81D': '短期融资券AA+',
             'CYCC41A': '同业存单AAA+', 'CYCC41B': '同业存单AAA', 'CYCC41C': '同业存单AA+',
             'CYCC80A': '企业债AAA+', 'CYCC80B': '企业债AAA', 'CYCC80D': '企业债AA+'}
dict_code = {'CYCC000': '国债', 'CYCC021': '政策性金融债（国开）',
             'CYCC82A': '中期票据AAA+', 'CYCC82B': '中期票据AAA', 'CYCC82C': '中期票据AAA-', 'CYCC82D': '中期票据AA+',
             'CYCC81A': '短期融资券AAA+', 'CYCC81B': '短期融资券AAA', 'CYCC81C': '短期融资券AAA-', 'CYCC81D': '短期融资券AA+',
             'CYCC41A': '同业存单AAA+', 'CYCC41B': '同业存单AAA', 'CYCC41C': '同业存单AA+',
             'CYCC80A': '企业债AAA+', 'CYCC80B': '企业债AAA', 'CYCC80D': '企业债AA+'}
step = ["0.25", "0.5", "0.75", "1.0", "1.5", "2.0"]

#国债 CYCC000 0.25 0.5 0.75 1.0 1.5 2.0
#http://www.chinamoney.com.cn/ags/ms/cm-u-bk-currency/ClsYldCurvHis?lang=CN&bondType=CYCC000&reference=1&startDate=2019-01-13&endDate=2019-01-20&termId=0.5
#政策性金融债（国开） CYCC021 0.25 0.5 0.75 1.0 1.5 2.0
#http://www.chinamoney.com.cn/dqs/rest/cm-u-bk-currency/ClsYldCurvHisExcel?lang=CN&bondType=CYCC021&reference=1&startDate=2019-01-13&endDate=2019-01-20&termId=0.5
#中期票据AA+(CYCC82D) AAA-(CYCC82C) AAA(CYCC82B) AAA+(CYCC82A) 0.25 0.5 0.75 1.0 1.5 2.0
#http://www.chinamoney.com.cn/dqs/rest/cm-u-bk-currency/ClsYldCurvHisExcel?lang=CN&bondType=CYCC82A&reference=1&startDate=2018-12-18&endDate=2019-01-18&termId=0.5
#短期融资券 AA+(CYCC81D) AAA-(CYCC81C) AAA(CYCC81B) AAA+(CYCC81A)
#企业债 AA+(CYCC80D) AAA(CYCC80B) AAA+(CYCC80A)
#同业存单 AA+(CYCC41C)  AAA(CYCC41B) AAA+(CYCC41A)

def cbk(a,b,c):
    '''回调函数
    @a:已经下载的数据块
    @b:数据块的大小
    @c:远程文件的大小
    '''
    per=100.0*a*b/c
    if per>100:
        per=100
    print('%.2f%%' % per)


def mkdir(path):
    folder = os.path.exists(path)

    if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
        print("---  Making folder /excel.  ---")
    else:
        print("---  Folder already exists.  ---")


def getshibor():
    global i_row
    # shibor def
    # {"head":{"version":"2.0","provider":"CWAP","req_code":"","rep_code":"200","rep_message":"","ts":1548040200660,"producer":""},"data":{"showDateEN":"21/01/2019 11:00","showDateCN":"2019-01-21 11:00"},"records":[{"termCode":"O/N","shibor":"2.2400","shibIdUpDown":"6.15","shibIdUpDownNum":6.1500000000,"termCodePath":"O/N"},{"termCode":"1W","shibor":"2.5940","shibIdUpDown":"2.40","shibIdUpDownNum":-2.4000000000,"termCodePath":"1W"},{"termCode":"2W","shibor":"2.6370","shibIdUpDown":"7.90","shibIdUpDownNum":7.9000000000,"termCodePath":"2W"},{"termCode":"1M","shibor":"2.7860","shibIdUpDown":"0.90","shibIdUpDownNum":0.9000000000,"termCodePath":"1M"},{"termCode":"3M","shibor":"2.9150","shibIdUpDown":"0.30","shibIdUpDownNum":-0.3000000000,"termCodePath":"3M"},{"termCode":"6M","shibor":"3.0550","shibIdUpDown":"0.60","shibIdUpDownNum":-0.6000000000,"termCodePath":"6M"},{"termCode":"9M","shibor":"3.2410","shibIdUpDown":"0.70","shibIdUpDownNum":-0.7000000000,"termCodePath":"9M"},{"termCode":"1Y","shibor":"3.3000","shibIdUpDown":"0.20","shibIdUpDownNum":-0.2000000000,"termCodePath":"1Y"}]}
    shibor_url = 'http://www.chinamoney.com.cn/r/cms/www/chinamoney/data/shibor/shibor.json'
    fp_html = requests.get(shibor_url).text
    fp_json = json.loads(fp_html)
    # print(fp_json['records'][0]["shibor"])
    for i in range(8):
        if i == 0:
            ws_out.cell(row=i_row, column=i_col).value = "Type"
        else:
            ws_out.cell(row=i_row, column=i_col + i).value = fp_json['records'][i-1]["termCode"]
    i_row += 1
    for i in range(8):
        if i == 0:
            ws_out.cell(row=i_row, column=i_col).value = "shibor"
        else:
            ws_out.cell(row=i_row, column=i_col + i).value = fp_json['records'][i-1]["shibor"]
    # print(i_row, i_col)
    i_row += 1
    print("Shibor数据拉取完成.")


def loadexcel(code):
    base_url = 'http://www.chinamoney.com.cn/dqs/rest/cm-u-bk-currency/ClsYldCurvHisExcel?lang=CN&bondType='
    sdate = '&reference=1&startDate='
    edate = '&endDate='
    end_url = '&termId=0.5'

    try:
        #python3
        if code == "Shibor":
            print('Begin download excel:' + all_code[code])
            # shibor_url = 'http://www.chinamoney.com.cn/dqs/rest/cm-u-bk-shibor/ShiborHisExcel?lang=cn&startDate=2018-12-23&endDate=2019-01-22'
            base_url = 'http://www.chinamoney.com.cn/dqs/rest/cm-u-bk-shibor/ShiborHisExcel?lang=cn&startDate='
            urllib.request.urlretrieve(base_url + days[0] + edate + days[6], homepath + code + '.xlsx')
        else:
            print('Begin download excel:' + all_code[code])
            urllib.request.urlretrieve(base_url + code + sdate + days[0] + edate + days[6] + end_url, homepath + os.sep + code + '.xlsx')
        #python2
        # urllib.urlretrieve(base_url + code + sdate + str(days[0]) + edate + str(days[6]) + end_url, homepath + code + '.xlsx')
    except Exception as e:
        print(e)


if __name__ == '__main__':
    mkdir(homepath)
    # getshibor()
    for item in all_code.keys():
        loadexcel(item)

    #把Shibor的数据拷贝进表格
    wb_shibor = load_workbook(homepath+ "Shibor" + '.xlsx')
    ws_shibor = wb_shibor.get_sheet_by_name("Sheet0")
    ws_out.cell(row=i_row, column=i_col).value = "Type"
    ws_out.cell(row=i_row, column=i_col+1).value = "日期"
    for i in range(8):
        ws_out.cell(row=i_row, column=i_col+i+2).value = shibor_step[i]
    i_row += 1
    for r in range(2, ws_shibor.max_row+1):
        if ws_shibor.cell(row=i_row, column=i_col).value in days:
            ws_out.cell(row=i_row, column=i_col).value = "Shibor"
            for s in range(9):
                ws_out.cell(row=i_row, column=i_col + s + 1).value = ws_shibor.cell(row=i_row, column=i_col + s).value
            i_row += 1
    print("Excel处理完成: Shibor")

    #开始拷贝债和票据的数据
    ws_out.cell(row=i_row, column=i_col).value = "债和票据"
    ws_out.cell(row=i_row, column=i_col+1).value = "日期"
    # 写入观察数据的步进值
    for i in range(6):
        # 空出首例类别和第二列的日期
        ws_out.cell(row=i_row, column=i_col+i+2).value = step[i]
    i_row += 1
    # 按序打开下载的excel，取数据出来
    for key in dict_code.keys():
        wb_tmp = load_workbook(homepath+ key + '.xlsx')
        ws_tmp = wb_tmp.get_sheet_by_name("Sheet0")
        # 填写类别，日期
        for d in range(7):
            ws_out.cell(row=i_row, column=i_col).value = dict_code[key]
            ws_out.cell(row=i_row, column=i_col+1).value = days[d]
            for r in range(1, ws_tmp.max_row + 1):
                if ws_tmp.cell(row=r, column=1).value == days[d] and ws_tmp.cell(row=r, column=2).value in step:
                    # 获取对应的季度在step里的索引，然后把数据表里第三列的值填入对应的索引表格
                    index = step.index(ws_tmp.cell(row=r, column=2).value)
                    ws_out.cell(row=i_row, column=i_col+2+index).value = ws_tmp.cell(row=r, column=3).value

            # 遍历完整个数据表，输出索引+1
            i_row += 1
        print("Excel处理完成:", dict_code[key])
    wb_out.save("Result.xlsx")
    print("数据输出完成，请打开Result.xlsx确认.")

