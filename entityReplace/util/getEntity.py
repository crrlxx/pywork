#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/5/17 15:30
# @Author  : crrlxx
# @contact: @gmail.com
# @Site    : 
# @File    : getEntity.py
# @Software: PyCharm Community Edition
from openpyxl import load_workbook
import re

wb_entity = load_workbook("../data/entitydata.xlsx")
ws_entity_data = wb_entity.get_sheet_by_name('entitydatasheet')
ws_entity_result = wb_entity.get_sheet_by_name('entityresultsheet')

# 存储每个domain、label组合下的entity内容
entity_list = {}
# 存储每个domain、label组合下的entity个数
count_domain_label = {}



def get_coding(str_input):
    '''
    获取编码格式
    '''
    if isinstance(str_input, unicode):
        return "unicode"
    try:
        str_input.decode("utf8")
        return 'utf8'
    except:
        pass
    try:
        str_input.decode("gbk")
        return 'gbk'
    except:
        pass


def find_rand_entity(domain, label, number):
    return entity_list[domain][label][number]


def sort_entity_list():
    total_row_num = ws_entity_data.max_row
    row_index = 1
    result_index = 1
    # 遍历所有实体，按语句类别分别存储实体类别和实体内容
    while row_index < total_row_num:
        # 获取实体标注列值
        cell_value = ws_entity_data.cell(row=row_index, column=2).value
        # 正则匹配出实体标签
        re_entity = re.compile(u'\<([a-z]+)\>([\u4e00-\u9fa5]+)\<\/[a-z]+\>')
        entity_type_test = re_entity.findall(cell_value)
        for i in entity_type_test:
            # 存储实体结果
            # 存储domain到ws_entity_result第一列
            ws_entity_result.cell(row=result_index, column=1).value = ws_entity_data.cell(row=row_index, column=1).value
            # 将实体标签存储到ws_entity_result第二列
            ws_entity_result.cell(row=result_index, column=2).value = i[0]
            # 将实体内容存储到ws_entity_result第三列
            ws_entity_result.cell(row=result_index, column=3).value = i[1]
            result_index += 1
        row_index += 1
        if row_index % 1000 == 0:
            print row_index, '/12000'
    wb_entity.save("../data/entitydata.xlsx")


def sort_label_entity():
    global entity_list
    global count_domain_label
    # 存储domain label entity
    domain_list = []
    label_list = {}
    label_list_sort = {}

    # 流程控制变量
    total_row_num = ws_entity_result.max_row
    row_index1 = 1
    row_index2 = 1
    row_index3 = 1

    # 分别查出所有的domain组合
    while row_index1 < total_row_num:
        domain_list.append(ws_entity_result.cell(row=row_index1, column=1).value)
        row_index1 +=1
    # 去重
    domain_list_sort = list(set(domain_list))
    # 根据domain的个数创建label list的字典
    for domain_index in domain_list_sort:
        label_list[domain_index] = []
    while row_index2 < total_row_num:
        # 把对应domain的label存到对应的字典里
        label_list[ws_entity_result.cell(row=row_index2, column=1).value].append(ws_entity_result.cell(row=row_index2, column=2).value)
        row_index2 += 1
    # 去重
    for domain_index in domain_list_sort:
        label_list_sort[domain_index] = list(set(label_list[domain_index]))
    print label_list_sort
    # 根据domain和label创建包含entity list的嵌套字典
    for domain_index in domain_list_sort:
        # 需要根据domain创建label的字典
        entity_list[domain_index] = {}
        count_domain_label[domain_index] = {}
        for label_index in label_list_sort[domain_index]:
            # 根据label创建entity的列表
            entity_list[domain_index][label_index] = []
            count_domain_label[domain_index][label_index] = 0
    while row_index3 < total_row_num:
        # 生成实体列表
        entity_list[ws_entity_result.cell(row=row_index3, column=1).value][ws_entity_result.cell(row=row_index3, column=2).value]\
            .append(ws_entity_result.cell(row=row_index3, column=3).value)
        row_index3 += 1
    # print entity_list
    # 存储个数
    for domain_index in domain_list_sort:
        for label_index in count_domain_label[domain_index]:
            count_domain_label[domain_index][label_index] = len(entity_list[domain_index][label_index])
    # print count_domain_label


def get_max_entity_count(domain, label):
    # 对传入参数做判断
    global count_domain_label
    if count_domain_label.has_key(domain):
        if count_domain_label[domain].has_key(label):
            return count_domain_label[domain][label]
        else:
            return 0
    else:
        print domain, label
        return 0


sort_label_entity()
