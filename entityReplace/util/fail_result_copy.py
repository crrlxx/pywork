#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/5/19 16:16
# @Author  : crrlxx
# @contact: @gmail.com
# @Site    : 
# @File    : fail_result_copy.py
# @Software: PyCharm Community Edition

from openpyxl import load_workbook

# 加载原始数据表
wb_failresult = load_workbook("../data/failresult.xlsx")
ws_failresult = wb_failresult.get_sheet_by_name('Sheet1')

total_row_number = ws_failresult.max_row
row_index = 2
while row_index < total_row_number:
    if ws_failresult.cell(row=row_index, column=3).value == 'F':
        ws_failresult.cell(row=row_index, column=6).value = ws_failresult.cell(row=row_index, column=1).value
    row_index += 1
wb_failresult.save("../data/failresult.xlsx")