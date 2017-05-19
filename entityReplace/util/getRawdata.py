#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2017/5/17 15:23
# @Author  : crrlxx
# @contact: @gmail.com
# @Site    : 
# @File    : getRawdata.py
# @Software: PyCharm Community Edition
from openpyxl import load_workbook
import re
import getEntity
import random


# 加载原始数据表
wb_rawdata = load_workbook("../data/aieresult.xlsx")
ws_rawdata = wb_rawdata.get_sheet_by_name('aieresult')
# 加载实体数据表
wb_entity = load_workbook("../data/entitydata.xlsx")
ws_entity_result = wb_entity.get_sheet_by_name('entityresultsheet')

# 从表格中逐行读取出domain及模板
total_row_num = ws_rawdata.max_row
row_index = 2
while row_index < total_row_num:
    raw_sentence = ws_rawdata.cell(row=row_index, column=1).value
    raw_domain = ws_rawdata.cell(row=row_index, column=2).value
    raw_entity = ws_rawdata.cell(row=row_index, column=4).value
    raw_model = ws_rawdata.cell(row=row_index, column=13).value
    # 待生成的新语句
    new_corpus = raw_model
    # 原始数据没有实体则直接拷贝原句子
    if raw_entity is None or raw_model is None:
        print row_index, 'no entity or no model'
        ws_rawdata.cell(row=row_index, column=14).value = raw_sentence
    else:
        # print '原始数据有实体'
        # 正则匹配出模板标签
        re_entity = re.compile(r'\[([a-z_]+)\]')
        try:
            entity_type = re_entity.findall(raw_model)
        except:
            print row_index,raw_model
        # 遍历匹配到的模版标签，在实体数据库里依次找到对应的实体填充标签
        for index in entity_type:
            # print '确认对应label个数'
            # 查询label对应的总数
            # 逻辑改动建议：提前在getEntity里统计得到对应的label及count，再做调用
            # 2017.05.19 已经在getEntity预先统计label和count
            label_entity_count = getEntity.get_max_entity_count(raw_domain, index)

            # 修改建议：将各个label的数据预先分类处理好，存储为结构体数据
            # 2017.05.19 已经预先处理数据
            # 如果label数不为0，则正常处理
            if label_entity_count != 0:
                # print '找到对应的实体label'
                # 在当前label的entity列表里随机取一个
                rand_num = random.randint(1, (label_entity_count-1))
                entity = getEntity.find_rand_entity(raw_domain, index, rand_num)
                # 替换原始模板标签
                new_corpus = new_corpus.replace(index, entity, 1)
            # 如果label为0，则从raw_entity里选择原标注数据填充
            else:
                print '没有找到，用原始数据填充'
                re_entity_orig = re.compile(u'\<([a-z_]+)\>([\u4e00-\u9fa5]+)\<\/[a-z_]+\>')
                entity_type_orig = re_entity_orig.findall(raw_entity)
                for inner in entity_type_orig:
                    print '找到原始数据中的实体内容'
                    print inner[0], index
                    if inner[0] == index:
                        new_corpus = new_corpus.replace(index, inner[1], 1)
        # 存储到最后一列
        # print '存储一条'
        ws_rawdata.cell(row=row_index, column=14).value = new_corpus
    row_index += 1
wb_rawdata.save("../data/aieresult.xlsx")

